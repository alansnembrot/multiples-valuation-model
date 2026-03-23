"""
Creates the base comparables.xlsx template.
Run this script ONCE to generate the file.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUTPUT_FILE = "Multiples Valuation Model with Yahoo Finance.xlsx"

C_DARK  = "1F3864"
C_MID   = "2F5496"
C_LIGHT = "D6E4F0"
C_INPUT = "EBF5FB"
C_GRAY  = "F2F2F2"

def fill(hex_c):
    return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def thin_border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="2F5496")
    return Border(left=s, right=s, top=s, bottom=s)

def h_font(size=10, color="FFFFFF", bold=True):
    return Font(name="Arial", size=size, color=color, bold=bold)

def b_font(size=10, color="000000", bold=False):
    return Font(name="Arial", size=size, color=color, bold=bold)


# ── INPUT SHEET ───────────────────────────────────────────────────────────────
def build_input_sheet(wb):
    ws = wb.active
    ws.title = "Input"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 3

    # Title
    ws.row_dimensions[1].height = 14
    ws.merge_cells("B2:D2")
    t = ws["B2"]
    t.value     = "MULTIPLES VALUATION MODEL — Yahoo Finance"
    t.font      = h_font(size=14)
    t.fill      = fill(C_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 40

    ws.row_dimensions[3].height = 10

    # TARGET section
    ws.merge_cells("B5:D5")
    h = ws["B5"]
    h.value     = "TARGET COMPANY"
    h.font      = h_font(size=10)
    h.fill      = fill(C_MID)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[5].height = 22

    ws["B6"].value     = "Ticker"
    ws["B6"].font      = b_font(bold=True)
    ws["B6"].fill      = fill(C_GRAY)
    ws["B6"].alignment = Alignment(vertical="center", indent=1)
    ws["B6"].border    = thin_border()
    ws.row_dimensions[6].height = 22

    ws["C6"].value     = "AAPL"
    ws["C6"].font      = Font(name="Arial", size=11, color="0000FF", bold=True)
    ws["C6"].fill      = fill(C_INPUT)
    ws["C6"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C6"].border    = thick_border()

    ws["D6"].value     = "← Enter a valid Yahoo Finance ticker"
    ws["D6"].font      = b_font(size=9, color="808080")
    ws["D6"].alignment = Alignment(vertical="center", indent=1)

    ws.row_dimensions[7].height = 10

    # COMPARABLE COMPANIES section
    ws.merge_cells("B8:D8")
    h2 = ws["B8"]
    h2.value     = "COMPARABLE COMPANIES"
    h2.font      = h_font(size=10)
    h2.fill      = fill(C_MID)
    h2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[8].height = 22

    comp_defaults = ["MSFT", "GOOGL", "META", "AMZN", "NVDA"]
    for i in range(5):
        row = 9 + i
        ws.row_dimensions[row].height = 22

        lc = ws.cell(row=row, column=2, value=f"Comparable {i+1}")
        lc.font      = b_font(bold=True)
        lc.fill      = fill(C_GRAY)
        lc.alignment = Alignment(vertical="center", indent=1)
        lc.border    = thin_border()

        ic = ws.cell(row=row, column=3, value=comp_defaults[i])
        ic.font      = Font(name="Arial", size=11, color="0000FF", bold=True)
        ic.fill      = fill(C_INPUT)
        ic.alignment = Alignment(horizontal="center", vertical="center")
        ic.border    = thick_border()

        hc = ws.cell(row=row, column=4, value=f"← Comparable {i+1}")
        hc.font      = b_font(size=9, color="808080")
        hc.alignment = Alignment(vertical="center", indent=1)

    ws.row_dimensions[14].height = 10

    # INSTRUCTIONS section
    ws.merge_cells("B15:D15")
    ih = ws["B15"]
    ih.value     = "INSTRUCTIONS"
    ih.font      = h_font(size=10)
    ih.fill      = fill(C_DARK)
    ih.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[15].height = 22

    instructions = [
        ("1.", "Enter tickers in the blue cells above"),
        ("2.", "Save this Excel file"),
        ("3.", "Open comparables_updater.py in Spyder"),
        ("4.", "Press F5 to run the script"),
        ("5.", "Results will appear in the Raw Data, Dashboard and Valuation tabs"),
        ("6.", "Enter the selected multiple in the Valuation tab after running the script"),
        ("",   ""),
        ("★",  "Blue text = user-editable input cells"),
        ("★",  "Tickers must be valid on Yahoo Finance (e.g. AAPL, MSFT, TSLA)"),
        ("★",  "See the Considerations tab for model assumptions and limitations"),
    ]

    for i, (num, text) in enumerate(instructions):
        row = 16 + i
        ws.row_dimensions[row].height = 18
        nc = ws.cell(row=row, column=2, value=num)
        nc.font      = b_font(bold=True, color="2F5496")
        nc.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(f"C{row}:D{row}")
        tc = ws.cell(row=row, column=3, value=text)
        tc.font      = b_font(size=9)
        tc.alignment = Alignment(vertical="center")




# ── RAW DATA SHEET (placeholder) ──────────────────────────────────────────────
def build_raw_data_sheet(wb):
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:V1")
    ph = ws["A1"]
    ph.value     = "This sheet is populated automatically when comparables_updater.py is run"
    ph.font      = b_font(size=11, color="808080")
    ph.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    ph.fill = fill(C_LIGHT)


# ── DASHBOARD SHEET (placeholder) ─────────────────────────────────────────────
def build_dashboard_sheet(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:V1")
    ph = ws["A1"]
    ph.value     = "This sheet is populated automatically when comparables_updater.py is run"
    ph.font      = b_font(size=11, color="808080")
    ph.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    ph.fill = fill(C_LIGHT)


# ── VALUATION SHEET (placeholder) ─────────────────────────────────────────────
def build_valuation_sheet(wb):
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:K1")
    ph = ws["A1"]
    ph.value     = "This sheet is populated automatically when comparables_updater.py is run"
    ph.font      = b_font(size=11, color="808080")
    ph.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    ph.fill = fill(C_LIGHT)


# ── CONSIDERATIONS SHEET ──────────────────────────────────────────────────────
def build_considerations_sheet(wb):
    ws = wb.create_sheet("Considerations")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 3

    C_WARN = "FCE4D6"   # light orange for limitations
    C_OK   = "E2EFDA"   # light green for considerations

    # Title
    ws.row_dimensions[1].height = 14
    ws.merge_cells("B2:C2")
    t = ws["B2"]
    t.value     = "MODEL CONSIDERATIONS & LIMITATIONS"
    t.font      = h_font(size=13)
    t.fill      = fill("1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:C3")
    sub = ws["B3"]
    sub.value     = "Read before using this model for investment decisions"
    sub.font      = b_font(size=9, color="595959", bold=True)
    sub.fill      = fill(C_LIGHT)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 10

    def section_header(row, text, color=C_DARK):
        ws.merge_cells(f"B{row}:C{row}")
        c = ws["B" + str(row)]
        c.value     = text
        c.font      = h_font(size=10)
        c.fill      = fill(color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    def item(row, label, text, bg=C_GRAY):
        lc = ws.cell(row=row, column=2, value=label)
        lc.font      = b_font(bold=True, size=9)
        lc.fill      = fill(bg)
        lc.alignment = Alignment(vertical="top", indent=1, wrap_text=True)
        lc.border    = thin_border()
        ws.row_dimensions[row].height = 40

        tc = ws.cell(row=row, column=3, value=text)
        tc.font      = b_font(size=9)
        tc.fill      = fill(bg)
        tc.alignment = Alignment(vertical="top", indent=1, wrap_text=True)
        tc.border    = thin_border()

    # MODEL SCOPE
    section_header(5, "MODEL SCOPE & SUPPORTED INDUSTRIES")
    item(6,  "Supported sectors",
         "Technology · Consumer Cyclical · Consumer Defensive · Healthcare / Pharma · "
         "Industrials / Manufacturing · Energy / Oil & Gas · Basic Materials · "
         "Communication Services. "
         "EV/EBITDA and EV/Revenue are meaningful and comparable within these sectors.")
    item(7,  "Not supported: Financial Services",
         "Banks, insurers, and diversified financials require P/BV and ROE-based frameworks. "
         "Their debt is a raw material (not leverage in the traditional sense), EBITDA is not "
         "meaningful, and Yahoo Finance data for these companies is frequently incomplete. "
         "A separate model is required.")
    item(8,  "Not supported: Real Estate / REITs",
         "REITs should be valued on FFO (Funds from Operations) and cap rates, not EV/EBITDA. "
         "Depreciation add-backs overstate EBITDA for asset-heavy real estate businesses. "
         "A REIT-specific model is required.")
    item(9,  "Utilities — supported with caveats",
         "EV/EBITDA applies but regulated utilities trade more on dividend yield and regulated "
         "asset base than on growth. Revenue and EBITDA growth metrics are less relevant. "
         "Use peer statistics with caution and prioritize EBITDA Margin and D/E.")

    ws.row_dimensions[10].height = 8

    # DATA SOURCE
    section_header(11, "DATA SOURCE")
    item(12, "Source",          "All financial data is sourced from Yahoo Finance via the yfinance Python library.")
    item(13, "Update frequency","Data is pulled at the time the script is run. It is NOT updated automatically. Re-run the script to refresh.")
    item(14, "Currency",        "All figures are reported in each company's trading currency (see 'Currency' column). No FX conversion is applied — do not compare companies in different currencies without adjusting.")
    item(15, "Data quality",    "Yahoo Finance data may contain errors, delays, or gaps (N/A). Always cross-check key figures against company filings (10-K, 10-Q) or Bloomberg/FactSet before use.")

    ws.row_dimensions[16].height = 8
    section_header(17, "FINANCIAL METRICS")
    item(18, "Revenue & EBITDA Growth", "Growth is calculated as a 2-year CAGR using actual dates from Yahoo Finance annual financials: CAGR = (t0 / t2)^(1/years) - 1, where t0 is the most recent valid data point and t2 is the valid point closest to 2 years prior by calendar date. Empty or TTM columns are skipped. The exact date range used is shown as a cell comment in the Raw Data and Dashboard sheets. Applied consistently across all companies regardless of fiscal year structure.")
    item(19, "EBITDA",          "EBITDA is taken directly from Yahoo Finance's TTM (trailing twelve months) field. It may differ from manually calculated EBITDA (EBIT + D&A) due to Yahoo's normalization methodology.")
    item(20, "Net Debt",        "Net Debt = Total Debt (balance sheet) minus Cash & Cash Equivalents. Pulled from the most recent balance sheet available. May not reflect intra-quarter changes.")
    item(21, "EBIT/Interest",   "Interest Coverage (EBIT/Interest Expense) is calculated manually from the income statement. Yahoo Finance does not provide this ratio directly.")

    ws.row_dimensions[23].height = 8
    section_header(24, "VALUATION METHODOLOGY")
    item(25, "Comparable selection",   "The quality of this analysis depends entirely on the quality of the comparable companies selected. Comparables should be in the same supported industry, similar in business model, geography, size, and growth profile. Avoid mixing companies from different sub-industries (e.g. software vs hardware) or significantly different market cap tiers, as these trade at structurally different multiples.")
    item(26, "Multiple selection",     "The suggested multiple is based on the target company's sector as reported by Yahoo Finance. The analyst should exercise judgment and may override this suggestion.")
    item(27, "Multiple range",         "The multiple range (min / base / max) is entered manually by the analyst based on the peer group statistics and their own judgment. This model does NOT automatically determine the range.")
    item(28, "Implied price formula",  "EV-based multiples: Implied Price = (Metric × Multiple − Net Debt) / Shares Outstanding. P/E-based: Implied Price = EPS × Multiple. Results are highly sensitive to the multiple selected.")
    item(29, "Not a recommendation",   "This model produces implied valuation ranges, NOT buy/sell recommendations. A full investment decision requires additional analysis including qualitative factors, risk assessment, and market context.")

    ws.row_dimensions[30].height = 8
    section_header(31, "KNOWN LIMITATIONS", color="C00000")
    item(32, "CAGR / LTM",     "Revenue and EBITDA growth is shown as 3-year CAGR. Absolute figures use the most recent fiscal year, not LTM. For companies with significant seasonality or recent M&A activity, this may introduce bias.")
    item(33, "No adjustments", "Figures are as-reported. No adjustments are made for one-time items, restructuring charges, or non-recurring revenues. Normalized EBITDA would require manual adjustment.")
    item(34, "Negative values","Companies with negative EBITDA, negative equity (P/BV), or negative earnings (P/E) will show N/A for the affected multiples. EV/Revenue may be more appropriate for these cases.")
    item(35, "Small peers",    "With fewer than 3 comparable companies, statistical measures (median, average) have limited reliability. Aim for at least 4-5 well-selected comparables.")
    item(36, "Currency mixing",  "Comparing companies in different currencies (e.g. USD vs EUR) without FX adjustment can distort size metrics (Market Cap, Revenue, EBITDA). If mixing currencies, treat absolute figures as indicative only and focus on ratios and margins.")
    item(37, "Size differences",  "Large-cap companies typically trade at a premium to small-cap peers (liquidity premium, analyst coverage, index inclusion). Where possible, keep peer market caps within the same order of magnitude as the target.")
    item(38, "Yahoo Finance",  "yfinance is an unofficial API. Yahoo Finance may change its data structure without notice, which can cause the script to fail or return incorrect data. Check for library updates regularly.")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    out_path   = os.path.join(script_dir, OUTPUT_FILE)

    wb = openpyxl.Workbook()
    build_input_sheet(wb)
    build_considerations_sheet(wb)
    build_raw_data_sheet(wb)
    build_dashboard_sheet(wb)
    build_valuation_sheet(wb)

    wb.save(out_path)
    print(f"✓ Template created: {out_path}")

if __name__ == "__main__":
    main()
