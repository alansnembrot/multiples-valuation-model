# =============================================================================
# MULTIPLES VALUATION MODEL — Data Extractor v5
# Engine: yfinance | Target: Multiples Valuation Model with Yahoo Finance.xlsx
#
# INSTRUCTIONS:
#   1. pip install yfinance openpyxl
#   2. Place this script in the same folder as the Excel file
#   3. Enter tickers in the Input sheet
#   4. Press F5 in Spyder
#
# METHODOLOGY NOTES:
#   - SUPPORTED SECTORS: Technology, Consumer (Cyclical & Defensive),
#     Healthcare, Industrials, Energy, Basic Materials, Communication Services.
#     Financial Services and Real Estate are NOT supported — their valuation
#     frameworks (P/BV, FFO, regulatory capital) require a different model.
#     Utilities are supported with caveats (growth metrics less relevant).
#   - Figures in each company's trading currency (see Currency column)
#   - Revenue & EBITDA Growth: 2-year CAGR calculated by finding the most recent
#     valid data point and the one closest to 2 years prior by actual date,
#     ignoring empty/TTM columns. Formula: (t0/t2)^(1/2) - 1.
#     Applied consistently across all companies regardless of fiscal year structure.
#   - EV/EBITDA and EV/Revenue: sourced directly from Yahoo Finance precalculated ratios
#     (enterpriseToEbitda, enterpriseToRevenue) — matches what Yahoo shows on the company page
#   - Net Debt = Total Debt (balance sheet) - Cash & equivalents
#   - EBIT/Interest Coverage calculated manually: EBIT / |Interest Expense|
#   - Peer group statistics exclude the target company
#   - Target company appears last in the data table (closest to peer statistics)
# =============================================================================

import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import statistics
import os
import warnings
from datetime import datetime
warnings.filterwarnings('ignore')

EXCEL_FILE = "Multiples Valuation Model with Yahoo Finance.xlsx"

# ── COLOR PALETTE ─────────────────────────────────────────────────────────────
COLOR_HEADER_DARK  = "1F3864"
COLOR_HEADER_MID   = "2F5496"
COLOR_ACCENT       = "D6E4F0"
COLOR_TARGET       = "FFF2CC"
COLOR_WHITE        = "FFFFFF"
COLOR_STATS_HEADER = "2E4057"
COLOR_STATS_ROW    = "EAF0FB"
COLOR_HIGHLIGHT    = "C6EFCE"
COLOR_HIGHLIGHT2   = "FFEB9C"
COLOR_GRAY_LIGHT   = "F2F2F2"
COLOR_INPUT_CELL   = "EBF5FB"

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
def h_font(white=True, size=10, bold=True):
    return Font(name="Arial", bold=bold, size=size,
                color="FFFFFF" if white else "000000")

def d_font(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def mk_fill(hex_c):
    return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def bdr(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bdr_thick():
    t = Side(style="medium", color="2F5496")
    return Border(left=t, right=t, top=t, bottom=t)

def ctr():
    return Alignment(horizontal="center", vertical="center")

def ctr_wrap():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_indent(n=1):
    return Alignment(horizontal="left", vertical="center", indent=n)


# ── DATA HELPERS ──────────────────────────────────────────────────────────────
def sf(val):
    try:
        f = float(val)
        return None if (f != f) else f
    except:
        return None

def to_millions(val):
    v = sf(val)
    return int(round(v / 1_000_000)) if v is not None else None

def to_ratio(val, dec=2):
    v = sf(val)
    return round(v, dec) if v is not None else None

def clean_name(raw_name):
    """
    Clean company name: strip trailing whitespace, collapse internal multiple
    spaces, remove common suffixes that create visual clutter (e.g. register marks).
    """
    if not raw_name:
        return raw_name
    import re
    name = raw_name.strip()
    name = re.sub(r'\s+', ' ', name)        # collapse multiple spaces
    name = re.sub(r'\s+[IVX]{1,4}$', '', name)  # strip trailing roman numerals
    name = name.strip()
    return name

def find_row(df, *keys):
    if df is None:
        return None
    for k in keys:
        if k in df.index:
            return df.loc[k]
    return None

def cagr_2y_by_date(row):
    """
    Calculate 2-year CAGR by finding valid data points by actual date.
    - Drops NaN/None values
    - Takes the most recent valid point as t0
    - Finds the valid point whose date is closest to 2 years (730 days) before t0
    - Returns (value, source_tag) where source_tag describes what was used
    
    This handles irregular fiscal years (TM, VOW.DE, etc.) where columns may
    be TTM, empty, or have non-standard spacing.
    """
    if row is None or len(row) == 0:
        return None, None

    import pandas as pd

    # Build list of (date, value) dropping nulls
    valid = []
    for date, val in row.items():
        v = sf(val)
        if v is not None:
            valid.append((pd.Timestamp(date), v))

    if len(valid) < 2:
        return None, None

    # Sort by date descending (most recent first)
    valid.sort(key=lambda x: x[0], reverse=True)

    t0_date, t0_val = valid[0]

    # Target date: 2 years before t0
    target_date = t0_date - pd.DateOffset(years=2)

    # Find the valid point closest to the target date
    best = min(valid[1:], key=lambda x: abs(x[0] - target_date))
    t2_date, t2_val = best

    # Calculate actual number of years between the two points
    years = (t0_date - t2_date).days / 365.25

    if years < 0.5:  # too close — not meaningful
        return None, None

    if t2_val <= 0 or t0_val <= 0:
        return None, None

    val = (t0_val / t2_val) ** (1.0 / years) - 1
    src = f"CAGR {t2_date.strftime('%b %y')}→{t0_date.strftime('%b %y')}"
    return val, src


# ── SECTOR CONFIG ─────────────────────────────────────────────────────────────
SECTOR_CONFIG = {
    "Technology": {
        "primary_multiple":    "EV/Revenue",
        "primary_vars":   ["Gross Margin", "Revenue Growth", "ROE"],
        "secondary_vars": ["EBITDA Margin", "EBITDA Growth", "ROA"],
    },

    "Consumer Cyclical": {
        "primary_multiple":    "EV/EBITDA",
        "primary_vars":   ["EBITDA Margin", "Revenue Growth"],
        "secondary_vars": ["Gross Margin", "ROE", "ROA"],
    },
    "Consumer Defensive": {
        "primary_multiple":    "EV/EBITDA",
        "primary_vars":   ["EBITDA Margin", "Revenue Growth"],
        "secondary_vars": ["Gross Margin", "ROE", "ROA"],
    },

    "Energy": {
        "primary_multiple":    "EV/EBITDA",
        "primary_vars":   ["EBITDA Margin", "Debt/Equity"],
        "secondary_vars": ["Revenue Growth", "ROE", "ROA"],
    },
    "Utilities": {
        "primary_multiple":    "EV/EBITDA",
        "primary_vars":   ["EBITDA Margin", "Debt/Equity"],
        "secondary_vars": ["Revenue Growth", "ROE", "ROA"],
    },
    "Healthcare": {
        "primary_multiple":    "EV/EBITDA",
        "primary_vars":   ["Gross Margin", "EBITDA Margin"],
        "secondary_vars": ["Revenue Growth", "ROE", "ROA"],
    },
    "Industrials": {
        "primary_multiple":    "EV/EBITDA",
        "primary_vars":   ["EBITDA Margin", "Revenue Growth"],
        "secondary_vars": ["ROE", "ROA", "Debt/Equity"],
    },
    "Basic Materials": {
        "primary_multiple":    "EV/EBITDA",
        "primary_vars":   ["EBITDA Margin", "Revenue Growth"],
        "secondary_vars": ["ROE", "ROA", "Debt/Equity"],
    },
    "Communication Services": {
        "primary_multiple":    "EV/EBITDA",
        "primary_vars":   ["EBITDA Margin", "Revenue Growth"],
        "secondary_vars": ["Gross Margin", "ROE", "ROA"],
    },
    "_default": {
        "primary_multiple":    "EV/EBITDA",
        "primary_vars":   ["EBITDA Margin", "Revenue Growth"],
        "secondary_vars": ["ROE", "ROA"],
    },
}

def get_sector_config(sector_str):
    return SECTOR_CONFIG.get(sector_str, SECTOR_CONFIG["_default"])


# ── DATA EXTRACTION ───────────────────────────────────────────────────────────
def fetch_metrics(ticker_str):
    print(f"  → Fetching {ticker_str}...")
    try:
        t    = yf.Ticker(ticker_str)
        info = t.info
        currency = info.get("currency", "USD")

        try:
            inc = t.financials
        except:
            inc = None

        # Revenue: 2Y CAGR by date
        rev_row              = find_row(inc, "Total Revenue", "Revenue")
        rev_growth, rev_src  = cagr_2y_by_date(rev_row)

        # EBITDA: 2Y CAGR by date
        ebitda_abs  = sf(info.get("ebitda"))
        ebitda_row  = find_row(inc, "EBITDA", "Normalized EBITDA")
        ebitda_curr = sf(ebitda_row.iloc[0]) if ebitda_row is not None and len(ebitda_row) > 0 else None
        ebitda_growth, ebi_src = cagr_2y_by_date(ebitda_row)

        ebitda_val = ebitda_abs or ebitda_curr   

        # EBIT / Interest Coverage
        ebit_row = find_row(inc, "EBIT", "Operating Income")
        ebit_val = sf(ebit_row.iloc[0]) if ebit_row is not None and len(ebit_row) > 0 else None

        int_row  = find_row(inc, "Interest Expense", "Interest Expense Non Operating")
        int_val  = sf(int_row.iloc[0]) if int_row is not None and len(int_row) > 0 else None
        if int_val is not None:
            int_val = abs(int_val)

        int_coverage = to_ratio(ebit_val / int_val, dec=1) \
                       if (ebit_val is not None and int_val and int_val != 0) else None

        # Balance Sheet
        try:
            bs = t.balance_sheet
        except:
            bs = None

        def bs_val(*keys):
            if bs is None:
                return None
            for k in keys:
                if k in bs.index:
                    v = sf(bs.loc[k].iloc[0])
                    if v is not None:
                        return v
            return None

        total_debt = (bs_val("Total Debt", "Long Term Debt And Capital Lease Obligation")
                      or sf(info.get("totalDebt")))
        cash       = (bs_val("Cash And Cash Equivalents",
                              "Cash Cash Equivalents And Short Term Investments",
                              "Cash And Short Term Investments")
                      or sf(info.get("totalCash")))

        net_debt = (total_debt - cash) if (total_debt is not None and cash is not None) else None

        nd_ebitda = to_ratio(net_debt / ebitda_val, dec=2) \
                    if (net_debt is not None and ebitda_val and ebitda_val != 0) else None

        de_raw  = sf(info.get("debtToEquity"))
        debt_eq = to_ratio(de_raw / 100, dec=2) if de_raw is not None else None

        market_cap  = sf(info.get("marketCap"))
        revenue_val = sf(info.get("totalRevenue"))

        # EV/EBITDA and EV/Revenue — Yahoo precalculated
        ev_ebitda  = to_ratio(info.get("enterpriseToEbitda"),  dec=1)
        ev_revenue = to_ratio(info.get("enterpriseToRevenue"), dec=1)

        # P/E — trailing preferred; forwardPE as fallback when EPS is negative
        pe        = sf(info.get("trailingPE"))
        pe_is_fwd = False
        if pe is None:
            fwd = sf(info.get("forwardPE"))
            if fwd is not None:
                pe        = to_ratio(fwd, dec=1)
                pe_is_fwd = True

        shares = sf(info.get("sharesOutstanding"))
        price  = sf(info.get("currentPrice") or info.get("regularMarketPrice"))
        eps    = sf(info.get("trailingEps"))

        return {
            "Ticker":          ticker_str,
            "Name":            clean_name(info.get("shortName", ticker_str)),
            "Sector":          info.get("sector", "N/A"),
            "Industry":        info.get("industry", "N/A"),
            "Currency":        currency,
            "Market Cap (M)":  to_millions(market_cap),
            "Revenue (M)":     to_millions(revenue_val),
            "EBITDA (M)":      to_millions(ebitda_val),
            "P/E":             pe,
            "_pe_is_fwd":      pe_is_fwd,
            "EV/EBITDA":       ev_ebitda,
            "EV/Revenue":      ev_revenue,
            "Gross Margin":    sf(info.get("grossMargins")),
            "EBITDA Margin":   sf(info.get("ebitdaMargins")),
            "ROE":             sf(info.get("returnOnEquity")),
            "ROA":             sf(info.get("returnOnAssets")),
            "Revenue Growth":  rev_growth,
            "EBITDA Growth":   ebitda_growth,
            "Net Debt (M)":    to_millions(net_debt),
            "Net Debt/EBITDA": nd_ebitda,
            "Debt/Equity":     debt_eq,
            "EBIT/Interest":   int_coverage,
            "_price":          price,
            "_eps":            eps,
            "_shares":         shares,
            "_net_debt_raw":   net_debt,
            "_revenue_raw":    revenue_val,
            "_ebitda_raw":     ebitda_val,
        }

    except Exception as e:
        print(f"  ✗ Error with {ticker_str}: {e}")
        return {"Ticker": ticker_str, "Name": f"ERROR: {e}"}


# ── COLUMN DEFINITIONS ────────────────────────────────────────────────────────
COLUMNS = [
    # Identification
    ("Ticker",          "Ticker",              "@",         10),
    ("Name",            "Company Name",        "@",         26),
    ("Sector",          "Sector",              "@",         20),
    ("Currency",        "Currency",            "@",          9),
    # Size & Scale
    ("Market Cap (M)",  "Market Cap (M)",      "#,##0",     14),
    ("Revenue (M)",     "Revenue (M)",         "#,##0",     14),
    ("EBITDA (M)",      "EBITDA (M)",          "#,##0",     13),
    # Valuation Multiples
    ("P/E",             "Price / Earnings",    "#,##0.0",   16),
    ("EV/EBITDA",       "EV / EBITDA",         "#,##0.0",   13),
    ("EV/Revenue",      "EV / Revenue",        "#,##0.0",   13),
    # Profitability
    ("Gross Margin",    "Gross Margin",        "0.0%",      13),
    ("EBITDA Margin",   "EBITDA Margin",       "0.0%",      13),
    ("ROE",             "Return on Equity",    "0.0%",      14),
    ("ROA",             "Return on Assets",    "0.0%",      14),
    # Growth
    ("Revenue Growth",  "Revenue Growth",      "0.0%",      14),
    ("EBITDA Growth",   "EBITDA Growth",       "0.0%",      13),
    # Financial Strength
    ("Net Debt (M)",    "Net Debt (M)",        "#,##0",     14),
    ("Net Debt/EBITDA", "Net Debt / EBITDA",   "#,##0.0",   14),
    ("Debt/Equity",     "Debt / Equity",       "#,##0.00",  13),
    ("EBIT/Interest",   "EBIT / Interest",     "#,##0.0",   13),
]

SECTIONS = [
    ("IDENTIFICATION",     4),
    ("SIZE & SCALE",       3),
    ("VALUATION MULTIPLES",3),
    ("PROFITABILITY",      4),
    ("GROWTH",             2),
    ("FINANCIAL STRENGTH", 4),
]

STAT_KEYS = [
    "Market Cap (M)", "Revenue (M)", "EBITDA (M)",
    "P/E", "EV/EBITDA", "EV/Revenue",
    "Gross Margin", "EBITDA Margin", "ROE", "ROA",
    "Revenue Growth", "EBITDA Growth",
    "Net Debt (M)", "Net Debt/EBITDA", "Debt/Equity", "EBIT/Interest",
]


# ── PEER GROUP STATISTICS ─────────────────────────────────────────────────────
def calc_peer_stats(metrics, target_ticker):
    comps = [m for m in metrics if m.get("Ticker") != target_ticker]
    stats = {}
    for key in STAT_KEYS:
        vals = [sf(m.get(key)) for m in comps]
        vals = [v for v in vals if v is not None]
        if vals:
            stats[key] = {
                "Median":  statistics.median(vals),
                "Average": statistics.mean(vals),
                "Min":     min(vals),
                "Max":     max(vals),
            }
        else:
            stats[key] = {"Median": None, "Average": None, "Min": None, "Max": None}
    return stats


# ── READ TICKERS FROM EXCEL ───────────────────────────────────────────────────
def read_tickers(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Input"]
    target = ws["C6"].value
    comps  = []
    for r in range(9, 14):
        v = ws.cell(row=r, column=3).value
        if v and str(v).strip():
            comps.append(str(v).strip().upper())
    wb.close()
    return (target.strip().upper() if target else None), comps


# ── MAIN TABLE WRITER ─────────────────────────────────────────────────────────
def write_table(ws, metrics, target, start_row, with_sections=True):
    """
    Writes comparables first, then target last — so target is adjacent
    to the peer statistics section below.
    """
    row = start_row

    if with_sections:
        ws.row_dimensions[row].height = 20
        col = 1
        for sec_name, sec_len in SECTIONS:
            end_col = col + sec_len - 1
            if col != end_col:
                ws.merge_cells(start_row=row, start_column=col,
                               end_row=row,   end_column=end_col)
            cell = ws.cell(row=row, column=col, value=sec_name)
            cell.font      = h_font(white=True, size=9)
            cell.fill      = mk_fill(COLOR_HEADER_MID)
            cell.alignment = ctr()
            cell.border    = bdr()
            col += sec_len
        row += 1

    # Column headers
    ws.row_dimensions[row].height = 28
    for ci, (_, label, fmt, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=row, column=ci, value=label)
        cell.font      = h_font(white=False, size=9, bold=True)
        cell.fill      = mk_fill(COLOR_ACCENT)
        cell.alignment = ctr_wrap()
        cell.border    = bdr()
        ws.column_dimensions[get_column_letter(ci)].width = width
    row += 1

    # Comparables first, then target last
    target_metric = next((m for m in metrics if m.get("Ticker") == target), None)
    comp_metrics  = [m for m in metrics if m.get("Ticker") != target]
    ordered = comp_metrics + ([target_metric] if target_metric else [])

    for m in ordered:
        is_tgt = (m.get("Ticker") == target)
        rf     = mk_fill(COLOR_TARGET) if is_tgt else mk_fill(COLOR_WHITE)
        ws.row_dimensions[row].height = 18
        for ci, (key, _, fmt, _) in enumerate(COLUMNS, 1):
            val  = m.get(key)
            disp = val if val is not None else "N/A"
            cell = ws.cell(row=row, column=ci, value=disp)
            cell.font      = d_font(bold=is_tgt)
            cell.fill      = rf
            cell.alignment = ctr()
            cell.border    = bdr()
            if val is not None and fmt != "@":
                cell.number_format = fmt
            # Mark forward P/E visually with '(fwd)' suffix
            if key == "P/E" and m.get("_pe_is_fwd") and val is not None:
                cell.value = f"{val:.1f} (fwd)"
                cell.number_format = "@"

        row += 1
    return row


# ── PEER STATS SECTION ────────────────────────────────────────────────────────
def write_peer_stats(ws, stats, start_row):
    total_cols  = len(COLUMNS)
    STATS_START = 4    # Column D
    row         = start_row

    ws.row_dimensions[row].height = 8
    row += 1

    # Section header from col D
    ws.merge_cells(start_row=row, start_column=STATS_START,
                   end_row=row,   end_column=total_cols)
    hc = ws.cell(row=row, column=STATS_START, value="Peer Group")
    hc.font      = h_font(white=True, size=10)
    hc.fill      = mk_fill(COLOR_STATS_HEADER)
    hc.alignment = left_indent(1)
    hc.border    = bdr()
    ws.row_dimensions[row].height = 22
    row += 1

    # Sub-header: col D = "Statistics", cols E+ = metric labels
    ws.row_dimensions[row].height = 24
    col_has_stats = {}
    for ci, (key, label, fmt, _) in enumerate(COLUMNS, 1):
        has = key in STAT_KEYS
        col_has_stats[ci] = has
        if ci < STATS_START:
            continue
        display_label = "Statistics" if ci == STATS_START else (label if has else "")
        cell = ws.cell(row=row, column=ci, value=display_label)
        cell.font      = h_font(white=False, size=9, bold=True)
        cell.fill      = mk_fill(COLOR_ACCENT)
        cell.alignment = ctr_wrap()
        cell.border    = bdr()
    row += 1

    # Stat rows — label in col D, values E+; A-C untouched
    for stat_label in ["Median", "Average", "Min", "Max"]:
        ws.row_dimensions[row].height = 18
        lc = ws.cell(row=row, column=STATS_START, value=stat_label)
        lc.font      = d_font(bold=True, size=9)
        lc.fill      = mk_fill(COLOR_STATS_ROW)
        lc.alignment = ctr()
        lc.border    = bdr()

        for ci, (key, _, fmt, _) in enumerate(COLUMNS, 1):
            if ci <= STATS_START:
                continue
            cell = ws.cell(row=row, column=ci)
            if col_has_stats[ci]:
                val  = stats.get(key, {}).get(stat_label)
                disp = val if val is not None else "N/A"
                cell.value         = disp
                cell.number_format = fmt if (val is not None and fmt != "@") else "@"
            else:
                cell.value = ""
            cell.font      = d_font(size=9)
            cell.fill      = mk_fill(COLOR_STATS_ROW)
            cell.alignment = ctr()
            cell.border    = bdr()
        row += 1

    return row


# ── RAW DATA SHEET ────────────────────────────────────────────────────────────
def write_raw_data(ws, metrics, target):
    ws.merged_cells.ranges.clear()
    ws.delete_rows(1, ws.max_row + 1)
    ws.sheet_view.showGridLines = False
    last_col = get_column_letter(len(COLUMNS))

    ws.merge_cells(f"A1:{last_col}1")
    tc = ws["A1"]
    tc.value     = "RAW DATA  |  Multiples Valuation Model  |  Source: Yahoo Finance"
    tc.font      = h_font(size=11)
    tc.fill      = mk_fill(COLOR_HEADER_DARK)
    tc.alignment = ctr()
    ws.row_dimensions[1].height = 28

    write_table(ws, metrics, target, start_row=2, with_sections=False)
    ws.freeze_panes = "A3"


# ── DASHBOARD SHEET ───────────────────────────────────────────────────────────
def write_dashboard(ws, metrics, target, peer_stats):
    ws.merged_cells.ranges.clear()
    ws.delete_rows(1, ws.max_row + 1)
    ws.sheet_view.showGridLines = False
    total_cols = len(COLUMNS)
    last_col   = get_column_letter(total_cols)

    ws.merge_cells(f"A1:{last_col}1")
    tc = ws["A1"]
    tc.value     = "MULTIPLES VALUATION MODEL — Comparable Companies Analysis"
    tc.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    tc.fill      = mk_fill(COLOR_HEADER_DARK)
    tc.alignment = ctr()
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A2:{last_col}2")
    nc = ws["A2"]
    nc.value = (
        "Source: Yahoo Finance  |  "
        "Peer group statistics exclude the target company"
    )
    nc.font      = Font(name="Arial", size=8, italic=True, color="595959")
    nc.fill      = mk_fill(COLOR_GRAY_LIGHT)
    nc.alignment = left_indent(1)
    ws.row_dimensions[2].height = 16

    # Row 3: Last updated timestamp
    ws.merge_cells(f"A3:{last_col}3")
    ts = ws["A3"]
    ts.value     = f"Last updated: {datetime.now().strftime('%d %b %Y  %H:%M')}"
    ts.font      = Font(name="Arial", size=8, italic=True, color="808080")
    ts.fill      = mk_fill(COLOR_GRAY_LIGHT)
    ts.alignment = left_indent(1)
    ws.row_dimensions[3].height = 14

    ws.row_dimensions[4].height = 6

    next_row = write_table(ws, metrics, target, start_row=5, with_sections=True)

    # Legend — target last so it's right above peer stats
    next_row += 1
    tgt_cell = ws.cell(row=next_row, column=1, value="★  Target Company")
    tgt_cell.font   = d_font(bold=True)
    tgt_cell.fill   = mk_fill(COLOR_TARGET)
    tgt_cell.border = bdr()
    ws.cell(row=next_row, column=2).fill   = mk_fill(COLOR_TARGET)
    ws.cell(row=next_row, column=2).border = bdr()
    ws.cell(row=next_row + 1, column=1,
            value="N/A = data not available on Yahoo Finance").font = d_font(size=8)

    next_row += 3
    write_peer_stats(ws, peer_stats, start_row=next_row)

    ws.freeze_panes = "A6"


# ── VALUATION VARS ────────────────────────────────────────────────────────────
VAL_VARS = [
    ("EBITDA Margin",   "EBITDA Margin",  "0.0%"),
    ("Gross Margin",    "Gross Margin",   "0.0%"),
    ("Revenue Growth",  "Revenue Growth", "0.0%"),
    ("EBITDA Growth",   "EBITDA Growth",  "0.0%"),
    ("ROA",             "ROA",            "0.0%"),
    ("ROE",             "ROE",            "0.0%"),
    ("Debt/Equity",     "Debt/Equity",    "#,##0.00"),
]


# ── VALUATION SHEET ───────────────────────────────────────────────────────────
def write_valuation(ws, metrics, target, sector_cfg, target_sector):
    ws.merged_cells.ranges.clear()
    ws.delete_rows(1, ws.max_row + 1)
    ws.sheet_view.showGridLines = False

    comps        = [m for m in metrics if m.get("Ticker") != target]
    multiplo_pri = sector_cfg["primary_multiple"]
    vars_pri     = sector_cfg["primary_vars"]
    vars_sec     = sector_cfg["secondary_vars"]
    target_data  = next((m for m in metrics if m.get("Ticker") == target), {})

    TARGET_COL  = 2
    MEDIAN_COL  = 3
    COMPS_START = 4
    title_end   = get_column_letter(max(8, COMPS_START + len(comps) + 1))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions[get_column_letter(TARGET_COL)].width = 16
    ws.column_dimensions[get_column_letter(MEDIAN_COL)].width = 14

    # Row 1: Title
    ws.merge_cells(f"A1:{title_end}1")
    tc = ws["A1"]
    tc.value = "Multiples Valuation"
    tc.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    tc.fill  = mk_fill(COLOR_HEADER_DARK)
    tc.alignment = ctr()
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 8

    # Rows 3-4: Industry / Multiple
    def label_value_row(r, label, value):
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = d_font(bold=True, size=10)
        lc.fill = mk_fill(COLOR_GRAY_LIGHT)
        lc.alignment = left_indent(1); lc.border = bdr()
        ws.row_dimensions[r].height = 22
        vc = ws.cell(row=r, column=2, value=value)
        vc.font = d_font(bold=True, size=10, color="1F3864")
        vc.fill = mk_fill(COLOR_ACCENT)
        vc.alignment = ctr(); vc.border = bdr()

    label_value_row(3, "Industry",          target_data.get("Industry", target_sector))
    label_value_row(4, "Selected Multiple", multiplo_pri)
    ws.row_dimensions[5].height = 8

    # Row 6: Table headers
    row = 6
    ws.row_dimensions[row].height = 26

    hc = ws.cell(row=row, column=1, value="Variable")
    hc.font = h_font(white=False, size=9, bold=True)
    hc.fill = mk_fill(COLOR_ACCENT); hc.alignment = ctr_wrap(); hc.border = bdr()

    th = ws.cell(row=row, column=TARGET_COL,
                 value=target_data.get("Ticker", "") + "\n" + target_data.get("Name", "")[:18])
    th.font = h_font(white=False, size=8, bold=True)
    th.fill = mk_fill(COLOR_TARGET); th.alignment = ctr_wrap(); th.border = bdr()

    mh = ws.cell(row=row, column=MEDIAN_COL, value="Median Peers")
    mh.font = h_font(white=True, size=8, bold=True)
    mh.fill = mk_fill(COLOR_STATS_HEADER); mh.alignment = ctr_wrap(); mh.border = bdr()

    for i, comp in enumerate(comps):
        ci = COMPS_START + i
        cell = ws.cell(row=row, column=ci,
                       value=comp["Ticker"] + "\n" + comp.get("Name", "")[:18])
        cell.font = h_font(white=False, size=8, bold=True)
        cell.fill = mk_fill(COLOR_ACCENT); cell.alignment = ctr_wrap(); cell.border = bdr()
        ws.column_dimensions[get_column_letter(ci)].width = 16
    row += 1

    # Variable rows
    def var_bg(k):
        if k in vars_pri:  return COLOR_HIGHLIGHT
        if k in vars_sec:  return COLOR_HIGHLIGHT2
        return COLOR_WHITE

    for var_label, var_key, var_fmt in VAL_VARS:
        bg = var_bg(var_key)
        ws.row_dimensions[row].height = 18

        lc = ws.cell(row=row, column=1, value=var_label)
        lc.font = d_font(bold=True, size=9)
        lc.fill = mk_fill(COLOR_WHITE)
        lc.alignment = left_indent(1); lc.border = bdr()

        tval = target_data.get(var_key)
        tc = ws.cell(row=row, column=TARGET_COL,
                     value=tval if tval is not None else "N/A")
        tc.font = d_font(bold=True, size=9)
        tc.fill = mk_fill(COLOR_TARGET); tc.alignment = ctr(); tc.border = bdr()
        if tval is not None and var_fmt != "@":
            tc.number_format = var_fmt

        comp_vals = [sf(c.get(var_key)) for c in comps if sf(c.get(var_key)) is not None]
        med_val   = statistics.median(comp_vals) if len(comp_vals) >= 2 \
                    else (comp_vals[0] if comp_vals else None)
        mc = ws.cell(row=row, column=MEDIAN_COL,
                     value=med_val if med_val is not None else "N/A")
        mc.font = d_font(bold=True, size=9, color="FFFFFF")
        mc.fill = mk_fill(COLOR_STATS_HEADER); mc.alignment = ctr(); mc.border = bdr()
        if med_val is not None and var_fmt != "@":
            mc.number_format = var_fmt

        for i, comp in enumerate(comps):
            ci  = COMPS_START + i
            val = comp.get(var_key)
            cell = ws.cell(row=row, column=ci,
                           value=val if val is not None else "N/A")
            cell.font = d_font(size=9, color="000000")
            cell.fill = mk_fill(COLOR_INPUT_CELL)
            cell.alignment = ctr(); cell.border = bdr()
            if val is not None and var_fmt != "@":
                cell.number_format = var_fmt
        row += 1

    # ── Valuation Results
    row += 2
    ws.row_dimensions[row].height = 8
    row += 1

    ebitda_raw   = target_data.get("_ebitda_raw")
    revenue_raw  = target_data.get("_revenue_raw")
    net_debt_raw = target_data.get("_net_debt_raw")
    shares_raw   = target_data.get("_shares")
    price_yahoo  = target_data.get("_price")
    eps_raw      = target_data.get("_eps")
    currency     = target_data.get("Currency", "USD")

    ev_based   = multiplo_pri in ("EV/EBITDA", "EV/Revenue")
    pe_based   = multiplo_pri == "P/E"
    metric_raw = ebitda_raw if multiplo_pri == "EV/EBITDA" else revenue_raw

    # ── Section header
    ws.merge_cells(f"A{row}:{title_end}{row}")
    rh = ws.cell(row=row, column=1, value="Valuation Results")
    rh.font = h_font(white=True, size=10)
    rh.fill = mk_fill(COLOR_STATS_HEADER)
    rh.alignment = left_indent(1)
    ws.row_dimensions[row].height = 22
    row += 1

    # ── Input rows: Yahoo values in col B (green), analyst override in col C (blue)
    # Formula hardcodes Yahoo values — if analyst overrides col C, they also
    # update the formula in the scenario rows manually (or via xlwings later)
    def input_row(r, label, auto_val, fmt, hint, units=""):
        ws.row_dimensions[r].height = 22
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = d_font(bold=True, size=10)
        lc.fill = mk_fill(COLOR_GRAY_LIGHT)
        lc.alignment = left_indent(1); lc.border = bdr()
        av = ws.cell(row=r, column=2, value=auto_val)
        av.font = d_font(bold=True, size=10, color="1F6128")
        av.fill = mk_fill(COLOR_GRAY_LIGHT)
        av.alignment = ctr(); av.border = bdr()
        av.number_format = fmt
        ov = ws.cell(row=r, column=3, value=None)
        ov.fill = mk_fill(COLOR_INPUT_CELL)
        ov.font = Font(name="Arial", size=10, color="0000FF", bold=True)
        ov.alignment = ctr(); ov.border = bdr_thick()
        ov.number_format = fmt
        hc = ws.cell(row=r, column=4, value=hint + (f"  [{units}]" if units else ""))
        hc.font = d_font(size=8, color="808080")
        hc.alignment = left_indent(1)

    ws.row_dimensions[row].height = 14
    for ci, lbl in enumerate(["Input", "Yahoo Finance (auto)", "Override (optional)", ""], 1):
        hc = ws.cell(row=row, column=ci, value=lbl)
        hc.font = d_font(bold=True, size=8, color="595959")
        hc.alignment = ctr()
    row += 1

    if ev_based:
        nd_v  = net_debt_raw if net_debt_raw is not None else 0
        sh_v  = shares_raw   if shares_raw   is not None else None
        met_v = metric_raw   if metric_raw   is not None else None
        met_lbl = "EBITDA (M)" if multiplo_pri == "EV/EBITDA" else "Revenue (M)"
        input_row(row, met_lbl,
                  (met_v / 1e6) if met_v else None, "#,##0.0",
                  f"Override {met_lbl} if you have adjusted figures", currency)
        row += 1
        input_row(row, "Net Debt (M)",
                  (nd_v / 1e6), "#,##0.0",
                  "Override Net Debt if balance sheet differs", currency)
        row += 1
        input_row(row, "Shares (M)",
                  (sh_v / 1e6) if sh_v else None, "#,##0.0",
                  "Override shares outstanding if diluted count differs", "M shares")
        row += 1
    elif pe_based:
        input_row(row, "EPS",
                  eps_raw, "#,##0.00",
                  "Override EPS if using normalized or forward estimate", currency)
        row += 1

    # ── Current price
    ws.row_dimensions[row].height = 10; row += 1
    ws.row_dimensions[row].height = 22
    pc_l = ws.cell(row=row, column=1, value="Current Price (Yahoo Finance)")
    pc_l.font = d_font(bold=True, size=10)
    pc_l.fill = mk_fill(COLOR_GRAY_LIGHT)
    pc_l.alignment = left_indent(1); pc_l.border = bdr()
    pc_v = ws.cell(row=row, column=2, value=price_yahoo)
    pc_v.font = d_font(bold=True, size=10, color="1F6128")
    pc_v.fill = mk_fill(COLOR_GRAY_LIGHT)
    pc_v.alignment = ctr(); pc_v.border = bdr()
    pc_v.number_format = "#,##0.00"
    price_row = row
    row += 1

    ws.row_dimensions[row].height = 10; row += 1

    # ── Scenario headers
    ws.row_dimensions[row].height = 20
    for ci, lbl in enumerate(["Scenario", "Multiple", "Implied Price", "Upside / Downside"], 1):
        hc = ws.cell(row=row, column=ci, value=lbl)
        hc.font = h_font(white=False, size=9, bold=True)
        hc.fill = mk_fill(COLOR_ACCENT)
        hc.alignment = ctr(); hc.border = bdr()
        ws.column_dimensions[get_column_letter(ci)].width = [20, 14, 18, 18][ci-1]
    row += 1

    SCENARIO_COLORS = {"Minimum": "FCE4D6", "Base Case": "E2EFDA", "Maximum": "DDEEFF"}

    for scenario in ["Minimum", "Base Case", "Maximum"]:
        sc_fill = mk_fill(SCENARIO_COLORS[scenario])
        ws.row_dimensions[row].height = 24

        sc = ws.cell(row=row, column=1, value=scenario)
        sc.font = d_font(bold=True, size=10)
        sc.fill = sc_fill; sc.alignment = ctr(); sc.border = bdr()

        mc = ws.cell(row=row, column=2, value=None)
        mc.fill = mk_fill(COLOR_INPUT_CELL)
        mc.font = Font(name="Arial", size=12, color="0000FF", bold=True)
        mc.alignment = ctr(); mc.border = bdr_thick()
        mc.number_format = "#,##0.0"
        mult_cell = f"B{row}"

        # Implied price — hardcoded Yahoo values
        # Override: analyst can manually update col C of input rows above
        if ev_based and metric_raw is not None and shares_raw:
            nd = net_debt_raw if net_debt_raw is not None else 0
            price_formula = f'=IF({mult_cell}="","N/A",({metric_raw}*{mult_cell}-{nd})/{shares_raw})'
        elif pe_based and eps_raw is not None:
            price_formula = f'=IF({mult_cell}="","N/A",{eps_raw}*{mult_cell})'
        else:
            price_formula = "N/A"

        pc = ws.cell(row=row, column=3, value=price_formula)
        pc.font = d_font(bold=True, size=11)
        pc.fill = sc_fill; pc.alignment = ctr(); pc.border = bdr()
        pc.number_format = "#,##0.00"

        ud_formula = f'=IF(OR({mult_cell}="",B{price_row}=0),"N/A",C{row}/B{price_row}-1)'
        ud = ws.cell(row=row, column=4, value=ud_formula)
        ud.font = d_font(bold=True, size=11)
        ud.fill = sc_fill; ud.alignment = ctr(); ud.border = bdr()
        ud.number_format = "+0.0%;-0.0%"

        row += 1

    row += 1
    ws.cell(row=row, column=1,
            value="Upside/Downside = Implied Price / Current Price (Yahoo) - 1  "
                  "|  To use adjusted figures, update the Override cells above "
                  "and re-enter the multiple in col B"
            ).font = d_font(size=8, color="595959")
    ws.row_dimensions[row].height = 14
    ws.freeze_panes = "A2"

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, EXCEL_FILE)

    if not os.path.exists(excel_path):
        print(f"✗ File not found: {excel_path}")
        return

    print("=" * 60)
    print("  MULTIPLES VALUATION MODEL — Data Extractor v5")
    print("=" * 60)

    print("\n[1/5] Reading tickers from Excel...")
    target, comps = read_tickers(excel_path)

    if not target:
        print("✗ Target ticker not found. Check cell C6 in the Input sheet.")
        return

    all_tickers = [target] + comps
    print(f"  Target:      {target}")
    print(f"  Comparables: {', '.join(comps) if comps else '(none)'}")

    print("\n[2/5] Fetching data from Yahoo Finance...")
    all_metrics = [fetch_metrics(t) for t in all_tickers]

    print("\n[3/5] Calculating peer group statistics...")
    peer_stats = calc_peer_stats(all_metrics, target)

    print("\n[4/5] Determining sector and valuation configuration...")
    target_data   = next((m for m in all_metrics if m.get("Ticker") == target), {})
    target_sector = target_data.get("Sector", "N/A")
    sector_cfg    = get_sector_config(target_sector)
    print(f"  Target sector:      {target_sector}")
    print(f"  Suggested multiple: {sector_cfg['primary_multiple']}")

    print("\n[5/5] Writing to Excel...")
    wb = openpyxl.load_workbook(excel_path)

    write_raw_data(wb["Raw Data"], all_metrics, target)
    write_dashboard(wb["Dashboard"], all_metrics, target, peer_stats)

    if "Valuation" in wb.sheetnames:
        del wb["Valuation"]
    wb.create_sheet("Valuation")
    write_valuation(wb["Valuation"], all_metrics, target, sector_cfg, target_sector)

    for old_name in ["Valuation Inputs"]:
        if old_name in wb.sheetnames:
            del wb[old_name]

    wb.save(excel_path)
    print(f"\n✓ Done  →  {excel_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
